#ExcelData
#Alexander Li
#2/17/16

from openpyxl import load_workbook
from aethalometer import *
from getters import *
from gps import *
import numpy
import geopy
from geopy.distance import vincenty
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
geolocator = Nominatim()
from peakFinder import *
from groupFinder import *
from scipy.signal import savgol_filter
import matplotlib.pyplot as plt
import time

#IMPORTANT: Things user changes

DATA_FILE = '16-01-27 AE33.xlsx' #File where data comes from
DESTINATION_FILE = 'fromCode2016.xlsx'
GPS_FILE = 'GPS Jan27-2016.xlsx' #File with GPS data

geoBool = False
individualRegressionBool = False
DCThreshold = 1500
twoSecondBool = False
smallSearch = -60 #for regression analysis of PM with DC
bigSearch = 60 #for regression analysis of PM with DC
peakSegmentation = False #To limit group sizes

globalShift = 0

print "Processing: " + DATA_FILE + ", " + GPS_FILE
print "Storing: " + DESTINATION_FILE
print "geoBool " + str(geoBool)
print "DC Threshold: " + str(DCThreshold)
print "Two Second Average: " + str(twoSecondBool)
print "Segmenting peaks: " + str(peakSegmentation)

def readExcel():
    """Loads an excel file, and finds the peaks from the DC column. Analyzes the
    peaks and prints the results in new excel file fromCode
    """
    #Data files
    wb = load_workbook(filename = DATA_FILE) #Load this data excel file
    ws = wb.active
    gpswb = load_workbook(filename = GPS_FILE) #Load this gps excel file
    gpsws = gpswb.active
    bottom = findBottom(ws) #Bottom of the data excel file
    bottomGPS = findBottom(gpsws) #Bottom of the gps excel file
    
    #Result files
    wb1 = load_workbook(filename = DESTINATION_FILE) #Load excel where you will print results
    ws1 = wb1.create_sheet("New Data") #Make new sheet
    
    #Gather lists
    DCList = getDCList(ws,bottom)
    timeList = getTimeList(ws,bottom)
    BC1 = getBC1List(ws,bottom)
    BC2 = getBC2List(ws,bottom)
    BC3 = getBC3List(ws,bottom)
    BC4 = getBC4List(ws,bottom)
    BC5 = getBC5List(ws,bottom)
    BC6 = getBC6List(ws,bottom)
    BC7 = getBC7List(ws,bottom)
    vList = getVList(gpsws, bottomGPS)
    llList = getLatLongList(gpsws, bottomGPS)
    pmList = getPMList(gpsws, bottomGPS)
    
    #Process lists
    bigGroupsList = getPeaksNew(ws1, DCList)
    #Change depending on fitting type
    
    #-----------------------------------------------------------------------------change
    result = calculateCurveFit(bigGroupsList, BC1, BC2, BC3, BC4, BC5, BC6, BC7, timeList)
    #-----------------------------------------------------------------------------change ^^^


    #Analyze the found peaks
    searchGroupsForPeaks(ws1, bigGroupsList, DCList)
    shift = performLinearRegression(ws1, bigGroupsList, DCList, pmList) #Shift removed <---------for wildfire
    #print "Shift: " + str(shift)
    
    #PutInExcelPortion
    putBCInExcel(ws1, BC6, BC7, bigGroupsList)
    putInExcel(ws1,result) #Put Aeth data in Excel
    putTimeInExcel(ws1, timeList, bigGroupsList) #Put timestamps in Excel
    putDCInExcel(ws1, bigGroupsList,DCList) #Put DC in excel
    putVInExcel(ws1, vList, bigGroupsList) #Put speeds in Excel
    putPMInExcel(ws1, pmList, bigGroupsList, shift) #Put PM2.5 in Excel <---------commented for wildfire
    putRangeInExcel(ws1, pmList, bigGroupsList, shift) #Put PM range in Excel <----commented for wildfire
    putLatLongAddressInExcel(ws1, llList, bigGroupsList) #Put LatLong in Excel
    wb1.save(DESTINATION_FILE)
    
    print 'SUCCESS!'

def getPeaksNew(excelResult, DCList):
    """Uses GroupFinder to find the peaks in the DC data"""
    groupFinder = GroupFinder(DCList)
    if peakSegmentation:
        groupFinder.segmentDC(DCThreshold)
    else:
        groupFinder.findElevatedDC(DCThreshold)
    for index in range(len(groupFinder.durationList)):
        excelResult.cell(row = index+2, column = 2, value = groupFinder.durationList[index])
    return groupFinder.elevatedDC

def searchGroupsForPeaks(excel, bigGroupsList, dCList):
    """Search through each bigGroup to see if there are individual peaks that are
    above the 1500 threshold
    CURRENTLY: Searching for inflection points indicating 2 peaks"""
    for index in range(len(bigGroupsList)):
        peakFinder = PeakFinder()
        peakFinder.locateInflectionPoint(bigGroupsList[index][0],bigGroupsList[index][1], dCList)
        excel.cell(row = index+2, column = 15, value = peakFinder.inflections)
   
   
        
def performLinearRegression(excel, bigGroupsList, dCList, pMList):
    """Perform linear regression to see how well the DC data and PM data correlate"""
    largestSum = 0
    largestShift = 0
    if individualRegressionBool:
        for index in range(len(bigGroupsList)):
            miniSum = 0
            miniShift = 0
            for shift in range(10):
                peakFinder = PeakFinder()
                peakFinder.performLinearRegression(bigGroupsList[index][0], bigGroupsList[index][1],dCList, pMList, shift)
                val = peakFinder.regressionScore 
                if val > miniSum: 
                    miniSum = val
                    miniShift = shift
            excel.cell(row = index+2, column = 16, value = cround(miniSum)) 
            excel.cell(row = index+2, column = 17, value = miniShift) 
    else:
        for shift in range(smallSearch,bigSearch):
            miniSum = 0
            for index in range(len(bigGroupsList)):
                peakFinder = PeakFinder()
                peakFinder.performLinearRegression(bigGroupsList[index][0], bigGroupsList[index][1],dCList, pMList, shift)
                miniSum += peakFinder.regressionScore#*(pMList[(bigGroupsList[index][0])])
            if miniSum > largestSum:
                largestSum = miniSum
                largestShift = shift
        for index in range(len(bigGroupsList)):
            miniSum = 0
            miniShift = 0
            for shift in range(largestShift-1,largestShift+2):
                peakFinder = PeakFinder()
                peakFinder.performLinearRegression(bigGroupsList[index][0], bigGroupsList[index][1],dCList, pMList, shift)
                val = peakFinder.regressionScore 
                if val > miniSum: 
                    miniSum = val
                    miniShift = shift
            excel.cell(row = index+2, column = 16, value = cround(miniSum))
            excel.cell(row = index+2, column = 17, value = miniShift)
        return largestShift
    
    
    
def calculate(bigGroupsList, BC1, BC2, BC3, BC4, BC5, BC6, BC7, timeList):
    result = []
    for x in bigGroupsList:
        #index = x[0] #obtain the index of the peaks
        if twoSecondBool:
            data = Data(averageListTwoSeconds(BC1,x[0],x[1]),averageListTwoSeconds(BC2,x[0],x[1]),averageListTwoSeconds(BC3,x[0],x[1]),
                    averageListTwoSeconds(BC4,x[0],x[1]),averageListTwoSeconds(BC5,x[0],x[1]),averageListTwoSeconds(BC6,x[0],x[1]),
                    averageListTwoSeconds(BC7,x[0],x[1]))
        else:
            data = Data(averageList(BC1,x[0],x[1]),averageList(BC2,x[0],x[1]),averageList(BC3,x[0],x[1]),
                    averageList(BC4,x[0],x[1]),averageList(BC5,x[0],x[1]),averageList(BC6,x[0],x[1]),
                    averageList(BC7,x[0],x[1]))
        data.analyze()
        data.variance()
        result.append(data.list)
        
        xVar = [370,470,520,590,660,880,950]
        yVar1 = data.BCdatalist
        yVar2 = data.ATN
        
        plt.plot(xVar,yVar1, 'b-', label='data')
        plt.plot(xVar, yVar2, 'r-', label='fit')
        print yVar2
        plt.xlabel('x')
        plt.ylabel('y')
        plt.legend()
        plt.show()
        
    return result

def averageList(list, start, finish):
    """Helper function for calculate.
    Input: The list and the indices you want to average through
    Averages the list values through the given indices."""
    count = finish-start
    accumulate = 0.0
    for index in range(start,finish):
        accumulate += list[index]
    average = accumulate/count
    return average

def averageListTwoSeconds(list, start, finish):
    """Input: The list and the indices you want to average
    Takes a 2 second average of the data. Groups every 2 points together.
    If there are an odd number of indices, remove the last index"""
    
    smoothed = savgol_filter(list[start:finish], 5, 2)
    return averageList(smoothed, 0, len(smoothed))
    """
    if (finish-start)%2 == 0: #even number of data in peak
        mini = []
        for x in range(start, finish, 2):
            mini.append(avg = (list[x] + list[x+1])/2)
        return averageList(mini, 0, len(mini))
    else:
        averageListTwoSeconds(list, start, finish-1)
    """

def putInExcel(excel, list):
    """Inputs the data into an excel file"""
    for i in range(len(list)):
        for j in range(len(list[0])):
            excel.cell(row = i+2,  column=j+3, value = list[i][j])

def putTimeInExcel(excel, timeList, groupsList):
    """Inputs the timestamps into the excel file"""
    times = getValuesFromIndices(timeList,groupsList)
    for i in range(len(times)):
        excel.cell(row = i+2, column = 1, value = str(times[i]))

def getValuesFromIndices(list,groupsList):
    """Takes the indices from bigGroupsList and uses it to create a list of the
    timestamps at those indices"""
    accum = []
    for i in range(len(groupsList)):
        mini = []
        for j in range(len(groupsList[0])):
            mini.append(list[groupsList[i][j]])
        accum.append(mini)
    return accum

def putDCInExcel(excel, groupsList, DCList):
    newDCList = []
    for indices in groupsList:
        newDCList.append(cround(averageList(DCList,indices[0],indices[1])))
    for i in range(len(newDCList)):
        excel.cell(row = i+2, column = 7, value = newDCList[i])
        
def putVInExcel(excel, vList, groupsList):
    """Inputs the average velocity of the timestamp into the excel file"""
    averageV=[]
    for i in groupsList:
        average = averageList(vList,i[0],i[1])
        averageV.append(cround(average))
    for i in range(len(averageV)):
        excel.cell(row = i+2, column = 8, value = averageV[i])
        
def putPMInExcel(excel, pmList, bigGroupsList, shift):
    """Inputs the average PM2.5 data of the timestamp into the excel file"""
    averagePM = []
    for i in bigGroupsList:
        average = averageList(pmList, i[0]-shift, i[1]-shift)
        averagePM.append(cround(average))
    for i in range(len(averagePM)):
        excel.cell(row = i+2, column = 13, value = averagePM[i])
        
def putRangeInExcel(excel, pmList, groupsList, shift):
    for x in range(len(groupsList)):
        peakFinder = PeakFinder()
        val = peakFinder.rangeFinder(groupsList[x][0]-shift, groupsList[x][1]-shift, pmList)
        excel.cell(row = x+2, column = 14, value = val)
        
def putLatLongAddressInExcel(excel, list, groupsList):
    """Puts the Latitude and Longitude data into the Excel File. Puts in 3 separate
    columns, the starting LatLong, the ending LatLong, and the distance traveled"""
    beginList = []
    endList = []
    distanceList = []
    addressList = []
    for i in groupsList:
        beginList.append(list[i[0]])
        endList.append(list[i[1]])
        distanceList.append(cround(vincenty(list[i[0]],list[i[1]]).meters))
    
    if geoBool:
        for coord in beginList:
            #location = geolocator.reverse(coord) 
            location = geocode(coord) #call geocode to handle timeouts
            address = location.address
            if address != None and 'Thompson Park, North Cayuga Street' in address:
                address = '120 Lake Ave'
            addressList.append(address)
       
    for i in range(len(beginList)):
        excel.cell(row = i+2, column = 10, value = str(beginList[i]))
        excel.cell(row = i+2, column = 11, value = str(endList[i]))
        excel.cell(row = i+2, column = 9, value = distanceList[i])
        if geoBool:
            excel.cell(row = i+2, column = 12, value = addressList[i]) #DON'T FORGET THIS IF MODIFYING GEOPY
            
def putBCInExcel(excel, BC6, BC7, groupsList):
    averageBC6andBC7 = []
    for i in range(len(groupsList)):
        excel.cell(row = i+2, column = 18, value = cround(averageList(BC6,groupsList[i][0],groupsList[i][1])))
        average = (averageList(BC6,groupsList[i][0],groupsList[i][1]) + averageList(BC7,groupsList[i][0],groupsList[i][1]))/2.0
        excel.cell(row = i+2, column = 19, value = cround(average))
            
def geocode(coord, recursion=0):
    try:
        return geolocator.reverse(coord)
    except GeocoderTimedOut as e:
        if recursion > 10:      # max recursions
            raise e

        time.sleep(1) # wait a bit
        # try again
        return geocode(coord, recursion=recursion + 1)
     
def findBottom(ws):
    """Finds the bottom index of the excel file. Tells you what value you should
    put for constant BOTTOM, and assigns BOTTOM.
    ***Note, the value put in for BOTTOM will be the read value-3"""
    CONTINUE = True
    read = 2
    while (CONTINUE):
        cell = ws['A'+str(read)]
        if (cell.value == None):
            CONTINUE = False
        else:
            read += 1
    return read-3
    
################################################################################

def cround(number):
        """Returns: the number rounded to 3 decimal places"""
        shift = number*10**3
        add = shift+0.5
        addint = int(add)
        addfloat = float(addint)
        shift2 = addfloat/10**3
        return shift2

################################################################################

def calculateCurveFit(bigGroupsList, BC1, BC2, BC3, BC4, BC5, BC6, BC7, timeList):
    result = []
    count = 0
    for x in bigGroupsList:
        #index = x[0] #obtain the index of the peaks
        if twoSecondBool:
            data = Data(averageListTwoSeconds(BC1,x[0],x[1]),averageListTwoSeconds(BC2,x[0],x[1]),averageListTwoSeconds(BC3,x[0],x[1]),
                    averageListTwoSeconds(BC4,x[0],x[1]),averageListTwoSeconds(BC5,x[0],x[1]),averageListTwoSeconds(BC6,x[0],x[1]),
                    averageListTwoSeconds(BC7,x[0],x[1]))
        else:
            data = Data(averageList(BC1,x[0],x[1]),averageList(BC2,x[0],x[1]),averageList(BC3,x[0],x[1]),
                    averageList(BC4,x[0],x[1]),averageList(BC5,x[0],x[1]),averageList(BC6,x[0],x[1]),
                    averageList(BC7,x[0],x[1]))
        #--------------------------------------------------
        data.curvefitWS() #Change for different types of Curve Fit!
        #--------------------------------------------------
        result.append(data.list)
        
        xVar = data.xdataAvg7 #MAKE SURE SAME AS curvefit
        yVar = data.ydataAvg7 #MAKE SURE SAME AS curvefit
        xWav = data.xWavelengthsAvg7 #MAKE SURE SAME AS curvefit
        
        #for comparing model with all measured data
        xWav7 = data.xWavelengths7 
        yVar7 = data.ydata7
        xVar7 = data.xdata7
        
        
        
        
        if count < 0: #change number of plots
            plt.plot(xWav,yVar, 'b-', label='data') #can make change to yVar here
            popt, pcov = optimize.curve_fit(data.func, xVar, yVar)
            plt.plot(xWav, data.func(xVar, *popt), 'r-', label='fit') #can make change to xVar here
            print yVar
            print "split"
            print data.func(xVar, *popt)
            plt.xlabel('wavelength')
            plt.ylabel('BC')
            plt.legend()
            plt.show()
            
            count += 1
        
    return result
    
################################################################################

readExcel()