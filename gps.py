#gps
#Alexander Li
#5/3/17

from openpyxl import load_workbook


SPEED = 'G'
START = '2' #Normally is 2
    
def getTList(ws,bottomGPS):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of timestamps"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    BList = []
    for row in ws.iter_rows(range_string='A'+START+':A' + str(bottomGPS)): #<------ Modify!
        for cell in row:
            full = str(cell.value)
            space = full.find(':')
            period = full.find('.')
            full = full[space-2:period]
            BList.append(full)
    return BList

def getVList(ws,bottomGPS):
    """Returns a list of all the velocities from the GPS data."""
    VList = []
    for row in ws.iter_rows(range_string= SPEED + START+ ':'+SPEED + str(bottomGPS)):
        for cell in row:
            if (cell.value == None):
                VList.append(0)
            else:
                VList.append(cround(float(cell.value)))
    return VList

def getLatLongList(ws, bottomGPS):
    LatList = getLatList(ws, bottomGPS)
    LongList = getLongList(ws, bottomGPS)
    lll = []
    for i in range(bottomGPS-int(START)):
        lll.append((LatList[i],LongList[i]))
    return lll

def getLatList(ws,bottomGPS):
    """Returns a list of all the Latitudes from the GPS data."""
    LatList = []
    for row in ws.iter_rows(range_string='B'+START+':B' + str(bottomGPS)):
        for cell in row:
            if (cell.value == None):
                LatList.append(0)
            else:
                LatList.append(float(cell.value))
    return LatList

def getLongList(ws,bottomGPS):
    """Returns a list of all the Latitudes from the GPS data."""
    LongList = []
    for row in ws.iter_rows(range_string='C'+START+':C' + str(bottomGPS)):
        for cell in row:
            if (cell.value == None):
                LongList.append(0)
            else:
                LongList.append(float(cell.value))
    return LongList

def getPMList(ws, bottomGPS):
    """Returns a list of all the PM2.5 from the GPS data"""
    pmList = []
    for row in ws.iter_rows(range_string = 'H' + START + ':H' + str(bottomGPS)):
        for cell in row:
            if (cell.value == None):
                pmList.append(0)
            else:
                pmList.append(float(cell.value))
    return pmList

def cround(number):
    """Returns: the number rounded to 3 decimal places"""
    shift = number*10**3
    add = shift+0.5
    addint = int(add)
    addfloat = float(addint)
    shift2 = addfloat/10**3
    return shift2
################################################################################
