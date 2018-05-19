#ExcelData
#Alexander Li
#2/17/16

from openpyxl import load_workbook
from aethalometer import *
from getters import *
from gps import *
import numpy
from scipy.signal import savgol_filter
import matplotlib.pyplot as plt
import time

def readExcel():
    """Loads an excel file, and finds the peaks from the DC column. Analyzes the
    peaks and prints the results in new excel file fromCode
    """
    #Data files
    wb = load_workbook(filename = "marchTioga2SecAvg.xlsx") #Load this data excel file
    ws = wb.active
    bottom = 10 #Bottom of the data excel file
    
    #Result files
    wb1 = load_workbook(filename = "marchTioga2SecAvg.xlsx") #Load excel where you will print results
    ws1 = wb1.create_sheet("New Data") #Make new sheet
    
    BC1 = getBC1List(ws,bottom)
    BC2 = getBC2List(ws,bottom)
    BC3 = getBC3List(ws,bottom)
    BC4 = getBC4List(ws,bottom)
    BC5 = getBC5List(ws,bottom)
    BC6 = getBC6List(ws,bottom)
    BC7 = getBC7List(ws,bottom)
    
    print BC1
    result = calculateCurveFit(BC1, BC2, BC3, BC4, BC5, BC6, BC7)
    print result
    
    putInExcel(ws1,result) #Put Aeth data in Excel
    
    wb1.save(filename = "marchTioga2SecAvg.xlsx")
    
def calculateCurveFit(BC1, BC2, BC3, BC4, BC5, BC6, BC7):
    result = []
    for x in range(len(BC1)):
        #index = x[0] #obtain the index of the peaks
        data = Data(BC1[x],BC2[x],BC3[x],BC4[x],BC5[x],BC6[x],BC7[x])
        data.curvefitAvg7() #Change for different types of Curve Fit!
        result.append(data.list)
        
        xVar = data.xdataAvg7 #MAKE SURE SAME AS curvefit
        xWav = data.xWavelengthsAvg7
        yVar = data.ydataAvg7 #MAKE SURE SAME AS curvefit
        
        plt.plot(xWav,yVar, 'b-', label='data')
        popt, pcov = optimize.curve_fit(data.func, xVar, yVar)
        plt.plot(xWav, data.func(xVar, *popt), 'r-', label='fit')
        plt.xlabel('wavelength')
        plt.ylabel('BC')
        plt.legend()
        plt.show()
        
    return result

def putInExcel(excel, list):
    """Inputs the data into an excel file"""
    for i in range(len(list)):
        for j in range(len(list[0])):
            excel.cell(row = i+2,  column=j+3, value = list[i][j])

readExcel()
    