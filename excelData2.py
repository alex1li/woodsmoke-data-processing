#excelData2
#Alexander Li
#4/27/16

from openpyxl import load_workbook
from aethalometer import *
import numpy

#IMPORTANT: Things user changes
BOTTOM = 7900 #Bottom index of the data in the excel file. Leave off bottom 2
DATA_FILE = '17-01-31 AE33.xlsx' #File where data comes from
DESTINATION_FILE = 'fromCode.xlsx'
CH1 = 'I'
CH2 = 'J'
CH3 = 'K'
CH4 = 'L'
CH5 = 'M'
CH6 = 'N'
CH7 = 'O'

def readExcel():
    """Loads an excel file, and finds the peaks from the DC column. Analyzes the
    peaks and prints the results in new excel file fromCode
    """
    wb = load_workbook(filename = DATA_FILE) #Load this excel file
    ws = wb.active
    DCList = getDCList(ws)
    timeList = getTimeList(ws)
    peaksList = getPeaks(DCList)
    groupsList = peakGroups(peaksList)
    bigGroupsList = removeSmallGroups(groupsList)
    print bigGroupsList
    BC1 = getBC1List(ws)
    BC2 = getBC2List(ws)
    BC3 = getBC3List(ws)
    BC4 = getBC4List(ws)
    BC5 = getBC5List(ws)
    BC6 = getBC6List(ws)
    BC7 = getBC7List(ws)
    result = calculate(bigGroupsList, BC1, BC2, BC3, BC4, BC5, BC6, BC7, timeList)
    wb1 = load_workbook(filename = DESTINATION_FILE) #Load excel where you will print results
    ws1 = wb1.create_sheet("New Data") #Make new sheet
    putInExcel(ws1,result) #Put Aeth data in Excel
    print result
    times = putTimeInExcel(ws1, timeList, bigGroupsList) #Put timestamps in Excel
    newDCList = putDCInExcel(ws1, bigGroupsList,DCList)
    wb1.save(DESTINATION_FILE)
    

def getPeaks(DCList):
    """Input: The list of DC values
    Return: The indices that are above 1500."""
    peaks = []
    for index in range(len(DCList)):
        if DCList[index] > 2000:
            peaks.append(index)
    return peaks

def peakGroups(peaksList):
    """Input: The list of peak indices from getPeaks
    Compare the index to the indices around it. Make sure the indices are adjacent.
    Then group adjacent indices. 
    Return: The peaks in groupings. Returns a 2D list"""
    groups = []
    minigroup = [peaksList[0]] #start with the first index
    singlePeaks = 0 #count the number of single instances >1500
    for index in range(len(peaksList)):
        if index < len(peaksList)-1: #if before end of the list
            if (peaksList[index+1]-peaksList[index] > 1) and (peaksList[index]-peaksList[index-1] > 1):
                singlePeaks += 1
            elif (peaksList[index+1]-peaksList[index] > 1): #end of a group
                minigroup.append(peaksList[index])
                groups.append(minigroup)
                minigroup = []
            elif (peaksList[index]-peaksList[index-1] > 1): #new group
                minigroup.append(peaksList[index])
        else:
            minigroup.append(peaksList[index]) #reach end of the list
            groups.append(minigroup)
            minigroup = []
            return groups

def removeSmallGroups(groupsList):
    """Input: The list of groups, with the indices of where the groups are
    Make sure the groups are 5 seconds or longer
    Return: The list of groups with small groups removed"""
    print groupsList
    for group in groupsList:
        if len(group) < 2:
            groupsList.remove(group)
            print 'caught an error in groupsList'
        elif group[1] - group[0] < 4:
            groupsList.remove(group)
    return groupsList


def calculate(bigGroupsList, BC1, BC2, BC3, BC4, BC5, BC6, BC7, timeList):
    result = []
    for x in bigGroupsList:
        #index = x[0] #obtain the index of the peaks
        data = Data(averageList(BC1,x[0],x[1]),averageList(BC2,x[0],x[1]),averageList(BC3,x[0],x[1]),
                    averageList(BC4,x[0],x[1]),averageList(BC5,x[0],x[1]),averageList(BC6,x[0],x[1]),
                    averageList(BC7,x[0],x[1]))
        data.curvefit()
        result.append(data.list)
    return result

def averageList(list, start,finish):
    """Helper function for calculate.
    Input: The list and the indices you want to average through
    Averages the list values through the given indices."""
    count = finish-start
    accumulate = 0.0
    for index in range(start,finish):
        accumulate += list[index]
    average = accumulate/count
    return average


def putInExcel(excel, list):
    """Inputs the data into an excel file"""
    for i in range(len(list)):
        for j in range(len(list[0])):
            excel.cell(row = i+2,  column=j+2, value = list[i][j])


def putTimeInExcel(excel, timeList, groupsList):
    """Inputs the timestamps into the excel file"""
    times = getTimes(timeList,groupsList)
    print times
    for i in range(len(times)):
        excel.cell(row = i+2, column = 1, value = str(times[i]))

def getTimes(timeList,groupsList):
    """Takes the indices from bigGroupsList and uses it to create a list of the
    timestamps at those indices"""
    accum = []
    for i in range(len(groupsList)):
        mini = []
        for j in range(len(groupsList[0])):
            mini.append(timeList[groupsList[i][j]])
        accum.append(mini)
    return accum

def putDCInExcel(excel, groupsList, DCList):
    newDCList = []
    for indices in groupsList:
        newDCList.append(cround(averageList(DCList,indices[0],indices[1])))
    for i in range(len(newDCList)):
        excel.cell(row = i+2, column = 6, value = newDCList[i])
    
################################################################################

def cround(number):
        """Returns: the number rounded to 3 decimal places"""
        shift = number*10**3
        add = shift+0.5
        addint = int(add)
        addfloat = float(addint)
        shift2 = addfloat/10**3
        return shift2
            
    
################################################################################
#Get List functions
    
def getDCList(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the DCList when given the 2 other columns in the excel file"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Hlist = []
    for row in ws.iter_rows(range_string=CH1+'2:'+CH1 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Hlist.append(int(cell.value))
    Mlist = []
    for row in ws.iter_rows(range_string=CH6+'2:'+CH6 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Mlist.append(int(cell.value))
    #Subtract Mlist values from Hlist values to obtain DC
    DClist=[]
    for x in range(len(Mlist)):
        DClist.append(Hlist[x]-Mlist[x])
    return DClist

def getTimeList(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of timestamps"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string='B2:B' + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(str(cell.value))
    return Blist

def getBC1List(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH1+'2:'+CH1 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(int(cell.value))
    return Blist

def getBC2List(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH2+'2:'+CH2 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(int(cell.value))
    return Blist

def getBC3List(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH3+'2:'+CH3 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(int(cell.value))
    return Blist

def getBC4List(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH4+'2:'+CH4 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(int(cell.value))
    return Blist

def getBC5List(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH5+'2:'+CH5 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(int(cell.value))
    return Blist

def getBC6List(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH6+'2:'+CH6 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(int(cell.value))
    return Blist

def getBC7List(ws):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH7+'2:'+CH7 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            Blist.append(int(cell.value))
    return Blist
    
readExcel()

################################################################################

"""   
def getPeaks(DClist):
    \"""Helper function.
    Input: The list of DC values
    Algortihm for determining peaks:
    (m1, m2, ...) < value < (n1, n2, ...)
    If the m values before the value are less than the value, and n values after
    the value are greater than the value, then the value is a peak. This allows
    us to eliminate tiny fluctuations in the data.
    NOTE! If you change m or n, also change averagePeaks()
    Returns: 2D list, with each node in the list containing [index, DC peak, timestamp]
    \"""
    #Algorithm to find peaks within DClist
    peaks = []
    for index in range(len(DClist)):
        if index >= 4 and index <= len(DClist)-5: #avoid edge cases
            n = [DClist[index+1],DClist[index+2]]
            m = [DClist[index-1],DClist[index-2]]
            if all(x < DClist[index] for x in n) & all(x < DClist[index] for x in m):
                if DClist[index] > 1500:  #make sure peak is of certain value
                    average = (n[0] + n[1] + m[1] + m[0])/4.0 #use average of surrounding values
                    #to ensure surrounding data doesn't vary too much
                    if 0 < (DClist[index]-average) < DClist[index]/10 : #make sure average isn't too different from value
                        peaks.append([index,DClist[index]]) #append 2D list with index of the value and everage of peaks
    return peaks
"""