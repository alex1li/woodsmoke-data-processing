#getters
#Alexander Li
#5/3/17

from openpyxl import load_workbook


#From the Excel File


CH1 = 'I'
CH2 = 'J'
CH3 = 'K'
CH4 = 'L'
CH5 = 'M'
CH6 = 'N'
CH7 = 'O'


def getDCList(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the DCList when given the 2 other columns in the excel file"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Hlist = []
    for row in ws.iter_rows(range_string=CH1+'2:'+CH1 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Hlist.append(0)
            else: 
                Hlist.append(int(cell.value))
    Mlist = []
    for row in ws.iter_rows(range_string=CH6+'2:'+CH6 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Mlist.append(0)
            else: 
                Mlist.append(int(cell.value))
    #Subtract Mlist values from Hlist values to obtain DC
    DClist=[]
    for x in range(len(Mlist)):
        DClist.append(Hlist[x]-Mlist[x])
    return DClist

def getTimeList(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of timestamps"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string='B2:B' + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(str(cell.value)) #May be int or str
    return Blist

def getBC1List(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH1+'2:'+CH1 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(int(cell.value))
    return Blist

def getBC2List(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH2+'2:'+CH2 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(int(cell.value))
    return Blist

def getBC3List(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH3+'2:'+CH3 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(int(cell.value))
    return Blist

def getBC4List(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH4+'2:'+CH4 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(int(cell.value))
    return Blist

def getBC5List(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH5+'2:'+CH5 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(int(cell.value))
    return Blist

def getBC6List(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH6+'2:'+CH6 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(int(cell.value))
    return Blist

def getBC7List(ws,BOTTOM):
    """Helper function. Note! Modify the range for different excel files! 
    Input: The excel file.
    Returns the list of BC1"""
    #Make sure the H and M columns match with the DC expression in the Excel sheet
    Blist = []
    for row in ws.iter_rows(range_string=CH7+'2:'+CH7 + str(BOTTOM)): #<------ Modify!
        for cell in row:
            if cell.value == None:
                Blist.append(0)
            else: 
                Blist.append(int(cell.value))
    return Blist

def findProperColumns(ws):
    """Finds the columns that that channels 1-7 are stored in the excel file"""
    letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']
    topRow = []
    for row in ws.iter_rows(range_string='A1:O1'):
        for cell in row:
            topRow.append(str(cell.value))
    ind = topRow.index('BC11')
    CH1 = topRow[ind]
    CH2 = topRow[ind+1]
    CH3 = topRow[ind+2]
    CH4 = topRow[ind+3]
    CH5 = topRow[ind+4]
    CH6 = topRow[ind+5]
    CH7 = topRow[ind+6]
            

        

    